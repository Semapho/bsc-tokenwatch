from bs4 import BeautifulSoup
from selenium import webdriver
from time import sleep
import os
from datetime import *

from pathlib import Path
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Here change and add path of your selenium diver in the variable path
PATH = r"enter path here"

# You can uncomment here the options, if you want chrome to be headless or invisible
options = webdriver.ChromeOptions()
options.add_argument("--enable-javascript")
#options.add_argument("--headless")
options.add_argument("user-agent=Mozilla/5.0 (Windows Phone 10.0; Android 4.2.1; Microsoft; Lumia 640 XL LTE) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Mobile Safari/537.36 Edge/12.10166")
options.add_argument('--log-level=3')

counter = 1
crypto_list = []
discarded = []               # list containing data of rugpulls crypto 
no_rugpulls = []            # list containing data of crypto which are not rugpulls
no_honeypots = []


def get_data(counter):
    dr = webdriver.Chrome(executable_path=PATH, chrome_options=options)
    dr.get("https://poocoin.app/ape")
    print("Waiting 25 sec for the content to load")
    sleep(25)
    bs = BeautifulSoup(dr.page_source, "html.parser")           #scraping poocoin.app/ape
    tbody = bs.tbody
    trs = tbody.contents
    for tr in trs:
        name, links = tr.contents[0:3:2]
        fixed_name = name.a.text

        for names in tr.contents[0::3]:
            name_link = (str(names)).split("=")

            for k in name_link:
                if k[0:8] == "\"/tokens":
                    namelinks = k

                else:
                    pass

        tryingfix_links = namelinks.split("\"")[1]
        fixed_name_links = "https://poocoin.app" + tryingfix_links

        for contracts in links.contents[0::2]:
            for bnbholders in links.contents[1::2]:
                cntr = (str(contracts)).split("=")        #trying to fix all names

                for i in cntr:
                    if i[0:5] == "\"http" and i[21] == "t":
                        contract_link = i[26:-9]
                    elif i[0:5] == "\"http" and i[21] == "a":
                        holder_link = i
                    else:
                        pass
                else:
                    pass
                fixed_contract_links = contract_link[1:-5]
                fixed_holder_links = holder_link[1:-5]

                liquidity = (str(bnbholders)).split("=")
                for j in liquidity:
                    if j[0:5] == "\"http":
                        bnbholder_link = j
                    else:
                        pass

                fixed_bnbholder_link = bnbholder_link[1:-5]

        crypto_name_links = fixed_name_links.splitlines()
        crypto_name = fixed_name.splitlines()
        crypto_contracts = fixed_contract_links.splitlines()
        crypto_holders = fixed_holder_links.splitlines()
        crypto_liquidity = fixed_bnbholder_link.splitlines()

        for table_name in crypto_name:
            for table_name_link in crypto_name_links:
                for table_contract in crypto_contracts:
                    for table_holders in crypto_holders:
                        for table_liquidity in crypto_liquidity:
                            tuplehere = (counter, table_name, table_name_link, table_contract, table_holders, table_liquidity)
                            crypto_list.append(tuplehere)
                            print("Found",counter,": ",tuplehere[1])
                            counter += 1                      #Storing data in a list
        


def check_rugpull(crypto_list,options,counter):
    
    for tuples in crypto_list:
        url = "https://bscscan.com/token/generic-tokenholders2?m=normal&a=" + str(tuples[3])
        #optionrug.add_argument("--headless")
        #optionrug.add_argument("--log-level=3")
        rugdr = webdriver.Chrome(PATH, chrome_options=options)
        rugdr.get(url)
        sleep(4)
        serialno = 0
        rug_soup = BeautifulSoup(rugdr.page_source, 'html.parser')
        
        rug_list = []
        
        try:
            itag = rug_soup.find_all(["i"], class_ = "far fa-file-alt text-secondary")
            span = itag[0].parent
            td = span.parent
            tr = td.parent
            one = tr.find("td")
            onec = one.text
    
        except:
            itag = None
            onec = 0
        
        for row in rug_soup.select("tr:has(td)"):
            tds = [td.get_text(strip=True) for td in row.select("td")]
            fix = tds[1].splitlines()
            for i in fix:
                rug_list.append(i)    

        try:
            first_address = rug_list[0][0:6]
            second_address = rug_list[1][0:6]
        
        except IndexError:
            second_address = None

        if ((first_address == "0x0000") or (first_address == "Burn A")) and (second_address == "Pancak"):
            no_rugpulls.append(tuples)
        
        elif (first_address == "Pancak") and ((second_address == "0x0000") or (second_address == "Burn A")):
            no_rugpulls.append(tuples)
        
        elif (first_address == "0x0000") or ((first_address == "Burn A") and (onec == "2")):
            no_rugpulls.append(tuples)
        
        elif ((first_address == "Pancak") or (first_address == "Burn A")) and (onec == "2"):
            no_rugpulls.append(tuples)
        
        elif (first_address == "Pancak"):
            no_rugpulls.append(tuples)
        
        else:
            discarded.append(tuples)

    print("Out of ",counter-1,", Non-Rugpulls are",len(no_rugpulls))
    print("They are:")
    
    for b in no_rugpulls:
        print(" ",b[1])

def check_honeypot(no_rugpulls,options):
    for data in no_rugpulls:
        
        honey_dr = webdriver.Chrome(PATH, chrome_options=options)
        honey_url = data[2]
        honey_dr.get(honey_url)
        sleep(7)

        honey_soup = BeautifulSoup(honey_dr.page_source,'html.parser')
        honey_dr.close()
        tags = honey_soup.find_all(text="Holders")
        pv2 = tags[0].parent
        link = pv2['href']
        bnblpaddress = link[26:-9]
        
        bnburl = "https://bscscan.com/token/generic-tokenholders2?m=normal&a=" + bnblpaddress + "&p=1"
        bnbdr = webdriver.Chrome(PATH)
        bnbdr.get(bnburl)
        sleep(4)

        bnbsoup = BeautifulSoup(bnbdr.page_source,'html.parser')
        bnbdr.close()
        honeylist = []
        
        try:
            itag = bnbsoup.find_all(["i"], class_ = "far fa-file-alt text-secondary")
            span = itag[0].parent
            td = span.parent
            tr = td.parent
            one = tr.find("td")
            onec = one.text
    
        except:
            itag = None
            onec = 0

        for rowsr in bnbsoup.select("tr:has(td)"):
            tdsr = [tdr.get_text(strip=True) for tdr in rowsr.select("td")]
        
            try:
                fixr = tdsr[1].splitlines()
                for r in fixr:
                    honeylist.append(r)
            except (IndexError, NameError) as e:
                pass
        
        try:    
            firstbnbaddress = honeylist[0][0:6]
        except IndexError:
            firstbnbaddress = "0000"

        if (firstbnbaddress == "0x0000") or (firstbnbaddress == "Burn A") or (firstbnbaddress[0:4] == "Legi"):
            no_honeypots.append(data)
    
        elif onec == "1":    
            no_honeypots.append(data)
        else:
            discarded.append(data)

    print("Out of Total",counter-1,", I chose after checking for Rugpulls and Honeypots",len(no_honeypots))
    print("They are:")
    for n in no_honeypots:
        print(" ",n[1])

def check_liquidity(no_honeypots,options):
    for values in no_honeypots:
        liq = webdriver.Chrome(executable_path=PATH, chrome_options=options)
        liq.get(values[2])
        sleep(5)
        liq_soup = BeautifulSoup(liq.page_source,'htmml.parser')
        try:
            span_tag = liq_soup.find_all(["span"], class_ = "text-success")
            liquid_value = span_tag[2].text
        except:
            liquid_value = "Null"
        
        last_tuple = values + (liquid_value,)
        print(last_tuple)

if __name__ == '__main__':
    
    get_data(counter)
    print()
    print("Now Checking for rugpulls and Honeypots")
    check_rugpull(crypto_list, options, counter)
    check_honeypot(no_rugpulls,options)
    check_liquidity(no_honeypots,options)
    print("Printing")

    directory = os.getcwd()
    todaysdate = str(date.today())
    path = "C:\\Users\\tusha\\Desktop\\python\\excel\\" + todaysdate
    my_file = Path(path)

    if my_file.is_file():
        
        time = str(datetime.now().time())
        fix_time = time.replace(':', '-')
        
        wb = load_workbook("excel\\" + todaysdate + ".xlsx")
        wb.create_sheet(title= fix_time, index=0)
        sheet = wb.active
        sheet['A1'] = "S.No."
        sheet['B1'] = "Name"
        sheet['C1'] = "Poocoin Link"
        sheet['D1'] = "Contract"
        sheet['E1'] = "Holders Link"
        sheet['F1'] = "BNB Holders Link"
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25
        
        for data in no_honeypots:
            sheet.append(data)
        for col in range(1,7):
            sheet[get_column_letter(col) + '1'].font = Font(bold=True)

        for data in discarded:
            sheet.append(data)
        for col in range(1,7):
            for row in range(len(no_honeypots)+2 , len(discarded) + len(discarded)):
                sheet[get_column_letter(col) + str(row)].font = Font(bold=True, color= "E2320B")
        
        sheet.insert_rows(len(no_honeypots)+2)
        sheet.insert_rows(len(no_honeypots)+2)
        
        wb.save(filename = "excel\\" + todaysdate + ".xlsx")



    
    else:
      
        time = str(datetime.now().time())
        fix_time = time.replace(':', '-')
        destinaton = "excel\\" + todaysdate + ".xlsx"
       
        wb = Workbook()
        wb.create_sheet(title= fix_time, index=0)
        sheet = wb.active
        sheet['A1'] = "S.No."
        sheet['B1'] = "Name"
        sheet['C1'] = "Poocoin Link"
        sheet['D1'] = "Contract"
        sheet['E1'] = "Holders Link"
        sheet['F1'] = "BNB Holders Link"

        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 25

        for data in no_honeypots:
            sheet.append(data)
        for col in range(1,7):
            sheet[get_column_letter(col) + '1'].font = Font(bold=True)

        for data in discarded:
            sheet.append(data)
        for col in range(1,7):
            for row in range(len(no_honeypots)+2 , len(discarded) + len(discarded)):
                sheet[get_column_letter(col) + str(row)].font = Font(bold=True, color= "E2320B")
        
        sheet.insert_rows(len(no_honeypots)+2)
        sheet.insert_rows(len(no_honeypots)+2)

        

        wb.save(filename = destinaton)

















